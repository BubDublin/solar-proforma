"""
⚡ Solar Pro-Forma Generator
A Streamlit web app for generating solar project pro-formas
"""

import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import io
from datetime import datetime

# Page configuration
st.set_page_config(
    page_title="Solar Pro-Forma Generator",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for styling
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1F4E78;
        text-align: center;
        margin-bottom: 0.5rem;
    }
    .sub-header {
        font-size: 1.2rem;
        color: #666;
        text-align: center;
        margin-bottom: 2rem;
    }
    .section-header {
        font-size: 1.3rem;
        font-weight: bold;
        color: #1F4E78;
        margin-top: 1.5rem;
        margin-bottom: 0.5rem;
        padding: 0.5rem;
        background-color: #E8F4F8;
        border-radius: 5px;
    }
    .metric-card {
        background-color: #F0F8FF;
        padding: 1rem;
        border-radius: 10px;
        border: 2px solid #1F4E78;
        text-align: center;
    }
    .metric-value {
        font-size: 1.8rem;
        font-weight: bold;
        color: #006100;
    }
    .metric-label {
        font-size: 0.9rem;
        color: #666;
    }
    .linked-value {
        background-color: #D9E1F2;
        padding: 0.3rem 0.5rem;
        border-radius: 3px;
        font-weight: bold;
        color: #305496;
    }
    .stButton>button {
        background-color: #1F4E78;
        color: white;
        font-size: 1.1rem;
        padding: 0.5rem 2rem;
        border-radius: 5px;
    }
    .stButton>button:hover {
        background-color: #2E5C8A;
    }
</style>
""", unsafe_allow_html=True)

# Header
st.markdown('<div class="main-header">⚡ Solar Pro-Forma Generator</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-header">Generate professional solar project pro-formas in seconds</div>', unsafe_allow_html=True)

# Initialize session state for defaults
if 'defaults_loaded' not in st.session_state:
    st.session_state.defaults_loaded = True
    # Default pricing from Notre Dame project
    st.session_state.modules_cost = 0.28
    st.session_state.inverters_cost = 0.15
    st.session_state.racking_cost = 0.23
    st.session_state.ballast_cost = 0.02
    st.session_state.electrical_cost = 0.23
    st.session_state.other_materials_cost = 0.02
    st.session_state.equipment_rental_cost = 0.04
    st.session_state.roof_attachments_cost = 0.02
    st.session_state.dumpsters_cost = 0.01
    st.session_state.porta_john_cost = 0.00
    st.session_state.safety_equipment_cost = 0.01
    st.session_state.engineering_cost = 0.02
    st.session_state.stamps_cost = 0.03
    st.session_state.permits_cost = 0.03
    st.session_state.meters_cost = 0.01
    st.session_state.ix_fees_cost = 0.00
    st.session_state.origination_cost = 0.05

# Sidebar for quick info
with st.sidebar:
    st.markdown("### 📊 About")
    st.info("""
    This tool generates professional solar project pro-formas 
    with live calculations for:
    • 25-year cash flow projections
    • SREC income (DC & MD markets)
    • Tax benefits (ITC & depreciation)
    • Payback analysis
    """)
    
    st.markdown("### 🔄 Current Market Data")
    st.markdown("**DC SREC (2025):** $380-455")
    st.markdown("**MD SREC (2025):** $48-50")
    st.markdown("**MD BT SREC:** $70-74")
    
    st.markdown("### 📅 Last Updated")
    st.markdown(f"{datetime.now().strftime('%B %d, %Y')}")

# Create tabs for organized input
st.markdown('<div class="section-header">📋 Project Information</div>', unsafe_allow_html=True)

col1, col2 = st.columns(2)

with col1:
    customer_name = st.text_input("Customer Name", value="NDMU", help="Enter the customer or organization name")
    project_name = st.text_input("Project Name", value="Notre Dame MD", help="Enter a name for this project")

with col2:
    system_size_kw = st.number_input("System Size (kW)", min_value=1.0, max_value=10000.0, value=236.6, step=0.1, help="Total system size in kilowatts")
    system_size_w = system_size_kw * 1000
    st.markdown(f"<span class='linked-value'>→ {system_size_w:,.0f} Watts</span>", unsafe_allow_html=True)

# Location
st.markdown('<div class="section-header">📍 Location & Utility</div>', unsafe_allow_html=True)

col1, col2 = st.columns(2)

with col1:
    jurisdiction = st.selectbox(
        "Jurisdiction",
        options=["Maryland", "Washington DC"],
        index=0,
        help="Select the project location for SREC calculations"
    )

with col2:
    utility = st.selectbox(
        "Utility Company",
        options=["PEPCO Maryland", "PEPCO DC", "BGE", "Potomac Edison", "Other"],
        index=0,
        help="Select the utility for electric rate defaults"
    )

# Set default electric rate based on utility
utility_rates = {
    "PEPCO Maryland": 0.135,
    "PEPCO DC": 0.147,
    "BGE": 0.110,
    "Potomac Edison": 0.125,
    "Other": 0.130
}

col1, col2, col3 = st.columns(3)

with col1:
    electric_rate = st.number_input(
        "Electric Rate ($/kWh)",
        min_value=0.05,
        max_value=0.50,
        value=utility_rates[utility],
        step=0.001,
        format="%.4f",
        help="Current electric rate per kWh"
    )

with col2:
    tsrf = st.number_input(
        "TSRF",
        min_value=500,
        max_value=2000,
        value=1250,
        step=10,
        help="Total Solar Resource Factor (kWh/kW/year)"
    )

with col3:
    tax_bracket = st.number_input(
        "Tax Bracket (%)",
        min_value=0.0,
        max_value=0.50,
        value=0.21,
        step=0.01,
        format="%.2f",
        help="Corporate tax bracket for depreciation benefits"
    )

# TOGGLES
st.markdown('<div class="section-header">⚙️ Toggle Settings</div>', unsafe_allow_html=True)

col1, col2, col3, col4 = st.columns(4)

with col1:
    itc_status = st.selectbox(
        "ITC Status",
        options=["With ITC (30%)", "Without ITC"],
        index=0,
        help="Federal Investment Tax Credit eligibility"
    )

with col2:
    if jurisdiction == "Maryland":
        srec_program = st.selectbox(
            "MD SREC Program",
            options=["Standard SREC", "Brighter Tomorrow SREC"],
            index=0,
            help="Maryland SREC program type"
        )
    else:
        srec_program = "Standard SREC (DC)"
        st.markdown("**SREC Program:** DC Standard")

with col3:
    escalation_rate = st.selectbox(
        "Escalation Rate",
        options=["2.0% (Conservative)", "3.5% (Moderate)", "5.0% (Aggressive)"],
        index=1,
        help="Annual utility rate increase assumption"
    )
    escalation = float(escalation_rate.split("%")[0]) / 100

with col4:
    degradation = st.number_input(
        "Panel Degradation (%/yr)",
        min_value=0.0,
        max_value=0.02,
        value=0.005,
        step=0.001,
        format="%.3f",
        help="Annual panel performance degradation"
    )

# PRICING INPUTS
st.markdown('<div class="section-header">💰 Pricing Inputs ($/Watt)</div>', unsafe_allow_html=True)

# Create pricing input columns
col1, col2, col3, col4 = st.columns(4)

pricing_inputs = {}

with col1:
    pricing_inputs['Modules'] = st.number_input("Modules", min_value=0.0, max_value=5.0, value=st.session_state.modules_cost, step=0.01, format="%.2f")
    pricing_inputs['Inverters'] = st.number_input("Inverters", min_value=0.0, max_value=5.0, value=st.session_state.inverters_cost, step=0.01, format="%.2f")
    pricing_inputs['Racking'] = st.number_input("Racking", min_value=0.0, max_value=5.0, value=st.session_state.racking_cost, step=0.01, format="%.2f")
    pricing_inputs['Ballast Block'] = st.number_input("Ballast Block", min_value=0.0, max_value=5.0, value=st.session_state.ballast_cost, step=0.01, format="%.2f")
    pricing_inputs['Electrical Material'] = st.number_input("Electrical Material", min_value=0.0, max_value=5.0, value=st.session_state.electrical_cost, step=0.01, format="%.2f")

with col2:
    pricing_inputs['Other Materials'] = st.number_input("Other Materials", min_value=0.0, max_value=5.0, value=st.session_state.other_materials_cost, step=0.01, format="%.2f")
    pricing_inputs['Equipment Rental'] = st.number_input("Equipment Rental", min_value=0.0, max_value=5.0, value=st.session_state.equipment_rental_cost, step=0.01, format="%.2f")
    pricing_inputs['Roof Attachments'] = st.number_input("Roof Attachments", min_value=0.0, max_value=5.0, value=st.session_state.roof_attachments_cost, step=0.01, format="%.2f")
    pricing_inputs['Dumpsters'] = st.number_input("Dumpsters", min_value=0.0, max_value=5.0, value=st.session_state.dumpsters_cost, step=0.01, format="%.2f")
    pricing_inputs['Porta John'] = st.number_input("Porta John", min_value=0.0, max_value=5.0, value=st.session_state.porta_john_cost, step=0.01, format="%.2f")

with col3:
    pricing_inputs['Safety Equipment'] = st.number_input("Safety Equipment", min_value=0.0, max_value=5.0, value=st.session_state.safety_equipment_cost, step=0.01, format="%.2f")
    pricing_inputs['Engineering'] = st.number_input("Engineering", min_value=0.0, max_value=5.0, value=st.session_state.engineering_cost, step=0.01, format="%.2f")
    pricing_inputs['Stamps'] = st.number_input("Stamps", min_value=0.0, max_value=5.0, value=st.session_state.stamps_cost, step=0.01, format="%.2f")
    pricing_inputs['Permits'] = st.number_input("Permits", min_value=0.0, max_value=5.0, value=st.session_state.permits_cost, step=0.01, format="%.2f")
    pricing_inputs['Revenue Grade Meters'] = st.number_input("Revenue Grade Meters", min_value=0.0, max_value=5.0, value=st.session_state.meters_cost, step=0.01, format="%.2f")

with col4:
    pricing_inputs['IX Application Fees'] = st.number_input("IX Application Fees", min_value=0.0, max_value=5.0, value=st.session_state.ix_fees_cost, step=0.01, format="%.2f")
    pricing_inputs['Origination Costs'] = st.number_input("Origination Costs", min_value=0.0, max_value=5.0, value=st.session_state.origination_cost, step=0.01, format="%.2f")

# Calculate totals
total_per_watt = sum(pricing_inputs.values())
total_cost = total_per_watt * system_size_w

# LIVE PREVIEW
st.markdown('<div class="section-header">📊 Live Preview</div>', unsafe_allow_html=True)

# Calculate key metrics
year1_production = system_size_kw * tsrf
year1_savings = year1_production * electric_rate

# SREC calculations
if jurisdiction == "Maryland":
    if srec_program == "Standard SREC":
        srec_value = 55 * 0.90  # ACP * 90%
    else:  # Brighter Tomorrow
        srec_value = 55 * 0.90 * 1.5  # ACP * 90% * 1.5x
else:  # DC
    srec_value = 460 * 0.85  # ACP * 85%

year1_srecs = year1_production / 1000
year1_srec_income = year1_srecs * srec_value

# ITC and depreciation
if itc_status == "With ITC (30%)":
    itc_amount = total_cost * 0.30
    depreciable_basis = total_cost * 0.85
else:
    itc_amount = 0
    depreciable_basis = total_cost

year1_depreciation = depreciable_basis  # 100% bonus in 2025
year1_depreciation_tax_savings = year1_depreciation * tax_bracket

after_itc_cost = total_cost - itc_amount
year1_total_benefit = year1_savings + year1_srec_income + year1_depreciation_tax_savings

# Simple payback (conservative - doesn't include all future years)
simple_payback = after_itc_cost / (year1_savings + year1_srec_income) if (year1_savings + year1_srec_income) > 0 else 0

# Display metrics
metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)

with metric_col1:
    st.markdown(f"""
    <div class="metric-card">
        <div class="metric-value">${total_cost:,.0f}</div>
        <div class="metric-label">Total System Cost</div>
    </div>
    """, unsafe_allow_html=True)

with metric_col2:
    st.markdown(f"""
    <div class="metric-card">
        <div class="metric-value">${itc_amount:,.0f}</div>
        <div class="metric-label">Federal Tax Credit</div>
    </div>
    """, unsafe_allow_html=True)

with metric_col3:
    st.markdown(f"""
    <div class="metric-card">
        <div class="metric-value">${year1_total_benefit:,.0f}</div>
        <div class="metric-label">Year 1 Total Benefit</div>
    </div>
    """, unsafe_allow_html=True)

with metric_col4:
    st.markdown(f"""
    <div class="metric-card">
        <div class="metric-value">{simple_payback:.1f} yrs</div>
        <div class="metric-label">Simple Payback</div>
    </div>
    """, unsafe_allow_html=True)

# Additional metrics
st.markdown("---")

col1, col2, col3 = st.columns(3)

with col1:
    st.markdown(f"**Cost per Watt:** ${total_per_watt:.2f}/W")
    st.markdown(f"**After-ITC Cost:** ${after_itc_cost:,.0f}")
    st.markdown(f"**Depreciable Basis:** ${depreciable_basis:,.0f}")

with col2:
    st.markdown(f"**Year 1 Production:** {year1_production:,.0f} kWh")
    st.markdown(f"**Year 1 Electric Savings:** ${year1_savings:,.0f}")
    st.markdown(f"**Year 1 SREC Income:** ${year1_srec_income:,.0f}")

with col3:
    st.markdown(f"**Year 1 SRECs:** {year1_srecs:.1f}")
    st.markdown(f"**SREC Value:** ${srec_value:.2f}")
    st.markdown(f"**Depr Tax Savings (Y1):** ${year1_depreciation_tax_savings:,.0f}")

# GENERATE BUTTON
st.markdown("---")

col1, col2, col3 = st.columns([1, 2, 1])

with col2:
    generate_button = st.button("📥 GENERATE PRO-FORMA EXCEL", use_container_width=True)

if generate_button:
    with st.spinner("Generating your pro-forma..."):
        # Create Excel workbook
        wb = Workbook()
        
        # Define styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        subheader_font = Font(bold=True, color="FFFFFF", size=11)
        input_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        currency_format = '$#,##0.00'
        pct_format = '0.00%'
        
        # TAB 1: Inputs & Assumptions
        ws_inputs = wb.active
        ws_inputs.title = "Inputs & Assumptions"
        
        # Set column widths
        ws_inputs.column_dimensions['A'].width = 3
        ws_inputs.column_dimensions['B'].width = 35
        ws_inputs.column_dimensions['C'].width = 20
        ws_inputs.column_dimensions['D'].width = 15
        ws_inputs.column_dimensions['E'].width = 30
        
        # Title
        ws_inputs['B1'] = f"{customer_name} - {project_name} - Pro-Forma Inputs"
        ws_inputs['B1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws_inputs['B1'].fill = header_fill
        ws_inputs.merge_cells('B1:E1')
        ws_inputs['B1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Toggle Settings
        row = 3
        ws_inputs[f'B{row}'] = "TOGGLE SETTINGS"
        ws_inputs[f'B{row}'].font = header_font
        ws_inputs[f'B{row}'].fill = header_fill
        ws_inputs.merge_cells(f'B{row}:E{row}')
        
        row += 1
        ws_inputs[f'B{row}'] = "Setting"
        ws_inputs[f'C{row}'] = "Value"
        ws_inputs[f'D{row}'] = "Unit"
        ws_inputs[f'E{row}'] = "Notes"
        for col in ['B', 'C', 'D', 'E']:
            ws_inputs[f'{col}{row}'].font = Font(bold=True)
            ws_inputs[f'{col}{row}'].fill = subheader_fill
            ws_inputs[f'{col}{row}'].font = subheader_font
        
        toggles = [
            ("ITC Status", itc_status, ""),
            ("SREC Program", srec_program, jurisdiction),
            ("Utility Company", utility, ""),
            ("Escalation Rate", f"{escalation:.1%}", ""),
        ]
        
        for label, value, note in toggles:
            row += 1
            ws_inputs[f'B{row}'] = label
            ws_inputs[f'C{row}'] = value
            ws_inputs[f'C{row}'].fill = input_fill
            ws_inputs[f'C{row}'].font = Font(bold=True)
            if note:
                ws_inputs[f'E{row}'] = note
        
        # Project Inputs
        row += 2
        ws_inputs[f'B{row}'] = "PROJECT INPUTS"
        ws_inputs[f'B{row}'].font = header_font
        ws_inputs[f'B{row}'].fill = header_fill
        ws_inputs.merge_cells(f'B{row}:E{row}')
        
        row += 1
        ws_inputs[f'B{row}'] = "Input"
        ws_inputs[f'C{row}'] = "Value"
        ws_inputs[f'D{row}'] = "Unit"
        for col in ['B', 'C', 'D']:
            ws_inputs[f'{col}{row}'].font = Font(bold=True)
            ws_inputs[f'{col}{row}'].fill = subheader_fill
            ws_inputs[f'{col}{row}'].font = subheader_font
        
        inputs = [
            ("Customer Name", customer_name, ""),
            ("Project Name", project_name, ""),
            ("System Size", system_size_kw, "kW"),
            ("Total System Cost", total_cost, "$"),
            ("Cost per Watt", total_per_watt, "$/W"),
            ("TSRF", tsrf, ""),
            ("Electric Rate", electric_rate, "$/kWh"),
            ("Tax Bracket", tax_bracket, "%"),
            ("Panel Degradation", degradation, "%/yr"),
        ]
        
        for label, value, unit in inputs:
            row += 1
            ws_inputs[f'B{row}'] = label
            ws_inputs[f'C{row}'] = value
            if isinstance(value, (int, float)) and value < 1 and value > 0:
                ws_inputs[f'C{row}'].number_format = pct_format
            elif isinstance(value, (int, float)) and value > 100:
                ws_inputs[f'C{row}'].number_format = currency_format
            ws_inputs[f'D{row}'] = unit
        
        # Cost Breakdown
        row += 2
        ws_inputs[f'B{row}'] = "COST BREAKDOWN"
        ws_inputs[f'B{row}'].font = header_font
        ws_inputs[f'B{row}'].fill = header_fill
        ws_inputs.merge_cells(f'B{row}:E{row}')
        
        row += 1
        ws_inputs[f'B{row}'] = "Category"
        ws_inputs[f'C{row}'] = "$/W"
        ws_inputs[f'D{row}'] = "Total ($)"
        for col in ['B', 'C', 'D']:
            ws_inputs[f'{col}{row}'].font = Font(bold=True)
            ws_inputs[f'{col}{row}'].fill = subheader_fill
            ws_inputs[f'{col}{row}'].font = subheader_font
        
        cost_start = row + 1
        for category, cost_per_w in pricing_inputs.items():
            row += 1
            ws_inputs[f'B{row}'] = category
            ws_inputs[f'C{row}'] = cost_per_w
            ws_inputs[f'C{row}'].number_format = '$0.00'
            ws_inputs[f'D{row}'] = cost_per_w * system_size_w
            ws_inputs[f'D{row}'].number_format = currency_format
        
        # Total
        row += 1
        ws_inputs[f'B{row}'] = "TOTAL PROJECT COST"
        ws_inputs[f'B{row}'].font = Font(bold=True)
        ws_inputs[f'C{row}'] = total_per_watt
        ws_inputs[f'C{row}'].font = Font(bold=True)
        ws_inputs[f'C{row}'].number_format = '$0.00'
        ws_inputs[f'D{row}'] = total_cost
        ws_inputs[f'D{row}'].font = Font(bold=True)
        ws_inputs[f'D{row}'].number_format = currency_format
        
        # TAB 2: 25-Year Cash Flow
        ws_cashflow = wb.create_sheet("25-Year Cash Flow")
        
        ws_cashflow.column_dimensions['A'].width = 3
        ws_cashflow.column_dimensions['B'].width = 8
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws_cashflow.column_dimensions[col].width = 14
        
        ws_cashflow['B1'] = f"{customer_name} - {project_name} - 25-Year Cash Flow"
        ws_cashflow['B1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws_cashflow['B1'].fill = header_fill
        ws_cashflow.merge_cells('B1:K1')
        ws_cashflow['B1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ["Year", "Electric Rate", "Production", "Electric Savings", "SREC Value", "SRECs", "SREC Income", "Tax Savings", "Total Benefit", "Cumulative"]
        for i, header in enumerate(headers):
            col = get_column_letter(i + 2)
            ws_cashflow[f'{col}3'] = header
            ws_cashflow[f'{col}3'].font = subheader_font
            ws_cashflow[f'{col}3'].fill = subheader_fill
            ws_cashflow[f'{col}3'].alignment = Alignment(horizontal='center', wrap_text=True)
        
        # ACP schedules
        md_acp = [55, 45, 35, 30, 30, 30, 30, 30, 0] + [0] * 16
        dc_acp = [460, 440, 420, 400, 380, 360, 340, 320, 300, 300, 300, 300, 300, 300, 300, 300, 300] + [100] * 8
        
        acp_schedule = dc_acp if jurisdiction == "Washington DC" else md_acp
        srec_multiplier = 1.5 if srec_program == "Brighter Tomorrow SREC" else 1.0
        srec_pct = 0.85 if jurisdiction == "Washington DC" else 0.90
        
        cumulative = -after_itc_cost
        
        for year in range(1, 26):
            row = 3 + year
            ws_cashflow[f'B{row}'] = year
            ws_cashflow[f'B{row}'].alignment = Alignment(horizontal='center')
            
            # Electric rate with escalation
            electric_rate_year = electric_rate * ((1 + escalation) ** (year - 1))
            ws_cashflow[f'C{row}'] = electric_rate_year
            ws_cashflow[f'C{row}'].number_format = '$0.0000'
            
            # Production with degradation
            production = year1_production * ((1 - degradation) ** (year - 1))
            ws_cashflow[f'D{row}'] = production
            ws_cashflow[f'D{row}'].number_format = '#,##0'
            
            # Electric savings
            savings = production * electric_rate_year
            ws_cashflow[f'E{row}'] = savings
            ws_cashflow[f'E{row}'].number_format = currency_format
            
            # SREC value
            acp = acp_schedule[year - 1] if year <= len(acp_schedule) else 0
            srec_val = acp * srec_pct * srec_multiplier
            ws_cashflow[f'F{row}'] = srec_val
            ws_cashflow[f'F{row}'].number_format = currency_format
            
            # SRECs
            srecs = production / 1000
            ws_cashflow[f'G{row}'] = srecs
            ws_cashflow[f'G{row}'].number_format = '#,##0.0'
            
            # SREC income
            srec_income = srecs * srec_val
            ws_cashflow[f'H{row}'] = srec_income
            ws_cashflow[f'H{row}'].number_format = currency_format
            
            # Tax savings (Year 1 only for bonus depreciation)
            tax_savings = year1_depreciation_tax_savings if year == 1 else 0
            ws_cashflow[f'I{row}'] = tax_savings
            ws_cashflow[f'I{row}'].number_format = currency_format
            
            # Total benefit
            total_benefit = savings + srec_income + tax_savings
            ws_cashflow[f'J{row}'] = total_benefit
            ws_cashflow[f'J{row}'].number_format = currency_format
            ws_cashflow[f'J{row}'].font = Font(bold=True)
            
            # Cumulative
            cumulative += total_benefit
            ws_cashflow[f'K{row}'] = cumulative
            ws_cashflow[f'K{row}'].number_format = currency_format
        
        # TAB 3: Client Summary
        ws_summary = wb.create_sheet("Client Summary")
        
        ws_summary.column_dimensions['A'].width = 3
        ws_summary.column_dimensions['B'].width = 35
        ws_summary.column_dimensions['C'].width = 20
        ws_summary.column_dimensions['D'].width = 15
        
        ws_summary['B1'] = f"{customer_name} - {project_name} - Client Summary"
        ws_summary['B1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws_summary['B1'].fill = header_fill
        ws_summary.merge_cells('B1:D1')
        ws_summary['B1'].alignment = Alignment(horizontal='center')
        
        # Project Overview
        row = 3
        ws_summary[f'B{row}'] = "PROJECT OVERVIEW"
        ws_summary[f'B{row}'].font = header_font
        ws_summary[f'B{row}'].fill = header_fill
        ws_summary.merge_cells(f'B{row}:D{row}')
        
        overview = [
            ("Customer", customer_name),
            ("Project", project_name),
            ("System Size", f"{system_size_kw} kW"),
            ("Total Cost", total_cost),
            ("Federal Tax Credit", itc_amount),
            ("After-ITC Cost", after_itc_cost),
        ]
        
        for label, value in overview:
            row += 1
            ws_summary[f'B{row}'] = label
            if isinstance(value, (int, float)) and value > 100:
                ws_summary[f'C{row}'] = value
                ws_summary[f'C{row}'].number_format = currency_format
            else:
                ws_summary[f'C{row}'] = value
        
        # Key Metrics
        row += 2
        ws_summary[f'B{row}'] = "KEY METRICS"
        ws_summary[f'B{row}'].font = header_font
        ws_summary[f'B{row}'].fill = header_fill
        ws_summary.merge_cells(f'B{row}:D{row}')
        
        metrics = [
            ("Year 1 Production", f"{year1_production:,.0f} kWh"),
            ("Year 1 Electric Savings", year1_savings),
            ("Year 1 SREC Income", year1_srec_income),
            ("Year 1 Total Benefit", year1_total_benefit),
            ("Simple Payback", f"{simple_payback:.1f} years"),
        ]
        
        for label, value in metrics:
            row += 1
            ws_summary[f'B{row}'] = label
            if isinstance(value, (int, float)) and value > 100:
                ws_summary[f'C{row}'] = value
                ws_summary[f'C{row}'].number_format = currency_format
            else:
                ws_summary[f'C{row}'] = value
        
        # 25-Year Totals
        row += 2
        ws_summary[f'B{row}'] = "25-YEAR TOTALS"
        ws_summary[f'B{row}'].font = header_font
        ws_summary[f'B{row}'].fill = header_fill
        ws_summary.merge_cells(f'B{row}:D{row}')
        
        # Calculate totals from cash flow
        total_electric_savings = sum([year1_production * electric_rate * ((1 + escalation) ** (y - 1)) * ((1 - degradation) ** (y - 1)) for y in range(1, 26)])
        total_srec_income = sum([(year1_production * ((1 - degradation) ** (y - 1)) / 1000) * (acp_schedule[y - 1] * srec_pct * srec_multiplier if y <= len(acp_schedule) else 0) for y in range(1, 26)])
        grand_total = total_electric_savings + total_srec_income + year1_depreciation_tax_savings
        
        totals = [
            ("Total Electric Savings", total_electric_savings),
            ("Total SREC Income", total_srec_income),
            ("Total Tax Benefits", year1_depreciation_tax_savings),
            ("GRAND TOTAL BENEFITS", grand_total),
        ]
        
        for i, (label, value) in enumerate(totals):
            row += 1
            ws_summary[f'B{row}'] = label
            ws_summary[f'C{row}'] = value
            ws_summary[f'C{row}'].number_format = currency_format
            if i == 3:  # Grand total
                ws_summary[f'B{row}'].font = Font(bold=True, size=12)
                ws_summary[f'C{row}'].font = Font(bold=True, size=12, color="006100")
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Offer download
        st.success("✅ Pro-forma generated successfully!")
        
        filename = f"{customer_name.replace(' ', '_')}_{project_name.replace(' ', '_')}_ProForma.xlsx"
        
        st.download_button(
            label="📥 Download Excel Pro-Forma",
            data=buffer,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

# Footer
st.markdown("---")
st.markdown("<div style='text-align: center; color: #666;'>Built for Captain Power Solar ⚡</div>", unsafe_allow_html=True)
